"""
Extract text từ file user upload — text/markdown/csv/pdf/xlsx.
Truncate output để fit context window (~50K chars).
"""
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_TEXT_CHARS = 50_000     # truncate text input cho LLM
MAX_FILE_BYTES = 10 * 1024 * 1024  # 10MB hard limit


SUPPORTED_EXT = {".txt", ".md", ".csv", ".tsv", ".pdf", ".xlsx", ".log", ".json"}


def is_supported(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_EXT


def extract_text(filename: str, file_bytes: bytes) -> tuple[str, str]:
    """Returns (text, summary_line). Summary = "PDF, 5 pages" hoặc "XLSX, 3 sheets" v.v.
    Raises ValueError nếu file lỗi / không hỗ trợ.
    """
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ValueError(f"File quá lớn ({len(file_bytes)/1024/1024:.1f}MB > 10MB)")

    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXT:
        raise ValueError(f"Không hỗ trợ extension {ext}. Hỗ trợ: {', '.join(sorted(SUPPORTED_EXT))}")

    if ext in (".txt", ".md", ".csv", ".tsv", ".log", ".json"):
        text = file_bytes.decode("utf-8", errors="replace")
        return _truncate(text), f"{ext.lstrip('.').upper()}, {len(text):,} chars"

    if ext == ".pdf":
        return _extract_pdf(file_bytes)

    if ext == ".xlsx":
        return _extract_xlsx(file_bytes)

    raise ValueError(f"Internal: ext {ext} not handled")


def _truncate(text: str) -> str:
    if len(text) <= MAX_TEXT_CHARS:
        return text
    return text[:MAX_TEXT_CHARS] + f"\n\n…(truncated, original {len(text):,} chars)"


def _extract_pdf(data: bytes) -> tuple[str, str]:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for i, page in enumerate(reader.pages):
        try:
            txt = page.extract_text() or ""
        except Exception as e:
            logger.warning(f"PDF page {i} extract error: {e}")
            txt = ""
        if txt.strip():
            pages.append(f"--- Page {i + 1} ---\n{txt.strip()}")
    text = "\n\n".join(pages)
    if not text.strip():
        raise ValueError("PDF không trích xuất được text — có thể là scan / image-only")
    return _truncate(text), f"PDF, {len(reader.pages)} pages"


def _extract_xlsx(data: bytes) -> tuple[str, str]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            sections.append(f"### Sheet: {sheet_name}\n\n" + "\n".join(rows))
    wb.close()
    text = "\n\n".join(sections)
    if not text.strip():
        raise ValueError("XLSX trống không có data")
    return _truncate(text), f"XLSX, {len(wb.sheetnames)} sheets"
